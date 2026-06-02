"""
AIDOC — Análisis Inteligente de Documentos
==========================================
Aplicación Streamlit completa en un solo archivo.

Requisitos (requirements.txt):
    streamlit
    langchain-community
    langchain-openai
    langchain-qdrant
    langchain-text-splitters
    qdrant-client
    pydantic-settings
    httpx
    pypdf
    openpyxl
    reportlab
    pandas

Ejecución:
    streamlit run aidoc_complete.py
"""

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — CONFIG (core/config.py)
# ══════════════════════════════════════════════════════════════════════════════

import urllib3
from functools import lru_cache
from pydantic_settings import BaseSettings, SettingsConfigDict

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class Settings(BaseSettings):
    model_config = SettingsConfigDict(env_file=".env", extra="ignore")

    LLM_URL: str = "https://llm-qwen3-30b-042026.project-pyxiia-proyectos.serving.ai-application.ciisagl.local/v1"
    LLM_API_KEY: str = "placeholder-llm-key"
    LLM_MODEL: str = "Qwen/Qwen3-30B-A3B-Instruct-2507-FP8"

    EMBEDDINGS_URL: str = "https://emb-qwen3-06b.project-user-jmartinez.serving.ai-application.ciisagl.local/v1"
    EMBEDDINGS_API_KEY: str = "placeholder-emb-key"
    EMBED_MODEL: str = "Qwen/Qwen3-Embedding-0.6B"

    QDRANT_URL: str = "https://qdrant.ai-application.ciisagl.local:443"
    QDRANT_COLLECTION: str = "aidoc_documents"

    LLM_TEMPERATURE: float = 0.1
    LLM_MAX_TOKENS: int = 1024
    EMBED_DIMENSIONS: int = 1024

    BATCH_WORKERS: int = 4
    CHUNK_SIZE: int = 2000
    CHUNK_OVERLAP: int = 200
    RETRIEVER_TOP_K: int = 8

    VERIFY_SSL: bool = False


@lru_cache
def get_settings() -> Settings:
    return Settings()


settings = get_settings()


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — INGESTION (core/ingestion.py)
# ══════════════════════════════════════════════════════════════════════════════
# Pipeline: PDF/TXT/ZIP → text extraction → chunking → embeddings → Qdrant

import zipfile
import tempfile
from pathlib import Path
from typing import Callable

import httpx
from langchain_community.document_loaders import PyPDFLoader, TextLoader
from langchain_openai import OpenAIEmbeddings
from langchain_qdrant import QdrantVectorStore
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, FieldCondition, Filter, MatchValue, VectorParams

ALLOWED_EXTENSIONS = {".pdf", ".txt", ".md"}


def _get_embeddings() -> OpenAIEmbeddings:
    return OpenAIEmbeddings(
        model=settings.EMBED_MODEL,
        base_url=settings.EMBEDDINGS_URL,
        api_key=settings.EMBEDDINGS_API_KEY,
        http_client=httpx.Client(verify=settings.VERIFY_SSL),
        http_async_client=httpx.AsyncClient(verify=settings.VERIFY_SSL),
    )


def _get_qdrant_client() -> QdrantClient:
    return QdrantClient(url=settings.QDRANT_URL, verify=settings.VERIFY_SSL)


def _get_store() -> QdrantVectorStore:
    return QdrantVectorStore(
        client=_get_qdrant_client(),
        collection_name=settings.QDRANT_COLLECTION,
        embedding=_get_embeddings(),
    )


def ensure_collection() -> None:
    client = _get_qdrant_client()
    existing = {c.name for c in client.get_collections().collections}

    if settings.QDRANT_COLLECTION not in existing:
        client.create_collection(
            collection_name=settings.QDRANT_COLLECTION,
            vectors_config=VectorParams(
                size=settings.EMBED_DIMENSIONS,
                distance=Distance.COSINE,
            ),
        )

    try:
        from qdrant_client.models import PayloadSchemaType
        client.create_payload_index(
            settings.QDRANT_COLLECTION,
            "metadata.filename",
            PayloadSchemaType.KEYWORD,
        )
    except Exception:
        pass


def ingest_file(file_path: str, filename: str, on_done: Callable | None = None) -> int:
    suffix = Path(file_path).suffix.lower()

    if suffix == ".pdf":
        loader = PyPDFLoader(file_path)
    else:
        loader = TextLoader(file_path, encoding="utf-8")

    docs = loader.load()
    for doc in docs:
        doc.metadata["filename"] = filename

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.CHUNK_SIZE,
        chunk_overlap=settings.CHUNK_OVERLAP,
    )
    chunks = splitter.split_documents(docs)

    store = _get_store()
    store.add_documents(chunks)

    if on_done:
        on_done(filename, len(chunks))

    return len(chunks)


def ingest_zip(zip_path: str, on_done: Callable | None = None) -> list[str]:
    ingested: list[str] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path, "r") as zf:
            for member in zf.namelist():
                p = Path(member)
                if p.suffix.lower() not in ALLOWED_EXTENSIONS:
                    continue
                if p.name.startswith(".") or "__MACOSX" in member:
                    continue
                zf.extract(member, tmpdir)
                extracted = Path(tmpdir) / member
                ingest_file(str(extracted), p.name, on_done)
                ingested.append(p.name)

    return ingested


def list_documents() -> list[str]:
    client = _get_qdrant_client()
    seen: set[str] = set()
    offset = None

    while True:
        records, offset = client.scroll(
            collection_name=settings.QDRANT_COLLECTION,
            limit=256,
            offset=offset,
            with_payload=["metadata"],
            with_vectors=False,
        )
        for r in records:
            fname = (r.payload or {}).get("metadata", {}).get("filename")
            if fname:
                seen.add(fname)
        if offset is None:
            break

    return sorted(seen)


def get_all_chunks_for_doc(filename: str) -> str:
    client = _get_qdrant_client()
    chunks: list[str] = []
    offset = None

    doc_filter = Filter(
        must=[FieldCondition(key="metadata.filename", match=MatchValue(value=filename))]
    )

    while True:
        records, offset = client.scroll(
            collection_name=settings.QDRANT_COLLECTION,
            scroll_filter=doc_filter,
            limit=100,
            offset=offset,
            with_payload=["page_content"],
            with_vectors=False,
        )
        for r in records:
            text = (r.payload or {}).get("page_content", "")
            if text:
                chunks.append(text)
        if offset is None:
            break

    return "\n\n".join(chunks)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — ANALYZER (core/analyzer.py)
# ══════════════════════════════════════════════════════════════════════════════
# Batch analyzer: N documents × M prompts → list of (filename, field_name, answer)
# Runs concurrently using asyncio + semaphore to respect BATCH_WORKERS limit.

import asyncio
from concurrent.futures import ThreadPoolExecutor

from langchain_openai import ChatOpenAI

SYSTEM_PROMPT = """\
Eres un analizador experto de documentos legales y contractuales.
Tu tarea es analizar el documento completo proporcionado y responder ÚNICAMENTE según la instrucción dada.

REGLAS:
- Basa tu respuesta SOLO en el texto del documento
- Si la información no está presente responde: "No encontrado"
- Respuestas concisas y estructuradas
- No inventes ni extrapoles información\
"""

ANALYSIS_TEMPLATE = """\
DOCUMENTO:
{document_text}

INSTRUCCIÓN:
{instruction}

RESPUESTA:\
"""


def _get_llm() -> ChatOpenAI:
    return ChatOpenAI(
        model=settings.LLM_MODEL,
        base_url=settings.LLM_URL,
        api_key=settings.LLM_API_KEY,
        temperature=settings.LLM_TEMPERATURE,
        max_tokens=settings.LLM_MAX_TOKENS,
        http_client=httpx.Client(verify=settings.VERIFY_SSL),
        http_async_client=httpx.AsyncClient(verify=settings.VERIFY_SSL),
    )


async def _analyze_pair(
    llm: ChatOpenAI,
    semaphore: asyncio.Semaphore,
    filename: str,
    prompt: dict,
) -> tuple[str, str, str]:
    async with semaphore:
        try:
            doc_text = await asyncio.to_thread(get_all_chunks_for_doc, filename)
            user_msg = ANALYSIS_TEMPLATE.format(
                document_text=doc_text,
                instruction=prompt["instruction"],
            )
            response = await llm.ainvoke([
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ])
            answer = response.content
        except Exception as exc:
            answer = f"Error: {exc}"

    return (filename, prompt["field_name"], answer)


async def _run_async(
    filenames: list[str],
    prompts: list[dict],
    on_progress: Callable | None,
) -> list[tuple]:
    llm = _get_llm()
    semaphore = asyncio.Semaphore(settings.BATCH_WORKERS)
    completed = 0
    total = len(filenames) * len(prompts)
    results: list[tuple] = []

    tasks = [
        _analyze_pair(llm, semaphore, fname, prompt)
        for fname in filenames
        for prompt in prompts
    ]

    for coro in asyncio.as_completed(tasks):
        result = await coro
        results.append(result)
        completed += 1
        if on_progress:
            on_progress(completed, total, result)

    return results


def run_batch_analysis(
    filenames: list[str],
    prompts: list[dict],
    on_progress: Callable | None = None,
) -> list[tuple[str, str, str]]:
    def _run():
        return asyncio.run(_run_async(filenames, prompts, on_progress))

    with ThreadPoolExecutor(max_workers=1) as pool:
        return pool.submit(_run).result()


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — CHATBOT (core/chatbot.py)
# ══════════════════════════════════════════════════════════════════════════════
# RAG chatbot over the analysis matrix + original documents in Qdrant.

import json

CHATBOT_SYSTEM_PROMPT_TEMPLATE = """\
Eres un asistente especializado en análisis documental (AIDOC).
Tienes acceso a:
1. La MATRIZ DE ANÁLISIS: resultados estructurados de {n_docs} documento(s) con {n_fields} campo(s) analizados por IA.
2. Los DOCUMENTOS ORIGINALES indexados en la base de conocimiento.

CAPACIDADES:
- Comparar resultados entre documentos
- Responder preguntas sobre campos específicos
- Profundizar en el texto original cuando sea necesario
- Generar tablas comparativas en markdown
- Alertar sobre inconsistencias o datos faltantes

{matrix_section}

INSTRUCCIÓN: Responde basándote primero en la matriz de análisis.
Si necesitas más detalle, usa el contexto adicional recuperado de los documentos originales.
Responde siempre en el mismo idioma que el usuario.\
"""

MATRIX_SECTION = """\
MATRIZ DE ANÁLISIS:
```json
{matrix_json}
```\
"""


def _build_matrix_context(results: list[tuple]) -> tuple[str, int, int]:
    if not results:
        return "", 0, 0

    matrix: dict[str, dict] = {}
    for fname, field, answer in results:
        matrix.setdefault(fname, {})[field] = answer

    n_docs = len(matrix)
    n_fields = max(len(v) for v in matrix.values()) if matrix else 0
    matrix_json = json.dumps(matrix, ensure_ascii=False, indent=2)
    return matrix_json, n_docs, n_fields


def _retrieve_context(query: str) -> str:
    try:
        store = QdrantVectorStore(
            client=QdrantClient(url=settings.QDRANT_URL, verify=settings.VERIFY_SSL),
            collection_name=settings.QDRANT_COLLECTION,
            embedding=OpenAIEmbeddings(
                model=settings.EMBED_MODEL,
                base_url=settings.EMBEDDINGS_URL,
                api_key=settings.EMBEDDINGS_API_KEY,
                http_client=httpx.Client(verify=settings.VERIFY_SSL),
                http_async_client=httpx.AsyncClient(verify=settings.VERIFY_SSL),
            ),
        )
        docs = store.similarity_search(query, k=settings.RETRIEVER_TOP_K)
        if not docs:
            return ""
        context = "\n\n---\n\n".join(
            f"[{d.metadata.get('filename', 'doc')} p.{d.metadata.get('page', '?')}]\n{d.page_content}"
            for d in docs
        )
        return f"\nCONTEXTO ADICIONAL DE DOCUMENTOS ORIGINALES:\n{context}"
    except Exception:
        return ""


def chat(
    messages: list[dict],
    results: list[tuple] | None = None,
) -> str:
    matrix_json, n_docs, n_fields = _build_matrix_context(results or [])

    matrix_section = (
        MATRIX_SECTION.format(matrix_json=matrix_json) if matrix_json else ""
    )

    system_prompt = CHATBOT_SYSTEM_PROMPT_TEMPLATE.format(
        n_docs=n_docs or "N/A",
        n_fields=n_fields or "N/A",
        matrix_section=matrix_section,
    )

    user_query = next(
        (m["content"] for m in reversed(messages) if m["role"] == "user"), ""
    )
    rag_context = _retrieve_context(user_query)

    llm_messages = [{"role": "system", "content": system_prompt}]
    for msg in messages[:-1]:
        llm_messages.append({"role": msg["role"], "content": msg["content"]})

    last_user = user_query + (f"\n\n{rag_context}" if rag_context else "")
    llm_messages.append({"role": "user", "content": last_user})

    llm = ChatOpenAI(
        model=settings.LLM_MODEL,
        base_url=settings.LLM_URL,
        api_key=settings.LLM_API_KEY,
        temperature=0.3,
        max_tokens=settings.LLM_MAX_TOKENS,
        http_client=httpx.Client(verify=settings.VERIFY_SSL),
        http_async_client=httpx.AsyncClient(verify=settings.VERIFY_SSL),
    )

    response = llm.invoke(llm_messages)
    return response.content


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — EXPORTERS (core/exporters.py)
# ══════════════════════════════════════════════════════════════════════════════
# Export analysis results to Excel (.xlsx) and PDF (.pdf)

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

BRAND_GREEN = "1E7145"
BRAND_GREEN_RGB = colors.HexColor(f"#{BRAND_GREEN}")
NOW = datetime.now()


def _pivot(results: list[tuple], prompts: list[dict]) -> tuple[list[str], dict]:
    fields = [p["field_name"] for p in prompts]
    matrix: dict[str, dict] = {}
    for fname, field, answer in results:
        matrix.setdefault(fname, {})[field] = answer
    return fields, matrix


def write_excel(results: list[tuple], prompts: list[dict]) -> bytes:
    fields, matrix = _pivot(results, prompts)
    wb = openpyxl.Workbook()

    _build_analysis_sheet(wb.active, fields, matrix)
    _build_metadata_sheet(wb.create_sheet("Metadata"), prompts)
    _build_prompts_sheet(wb.create_sheet("Prompts Used"), prompts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_analysis_sheet(ws, fields: list[str], matrix: dict) -> None:
    ws.title = "Analysis Results"
    headers = ["Filename"] + fields

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor=BRAND_GREEN)
    header_align = Alignment(wrap_text=True, vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = 35

    ws.row_dimensions[1].height = 24

    for row_i, (fname, fields_map) in enumerate(matrix.items(), start=2):
        alt = row_i % 2 == 0
        fill = PatternFill("solid", fgColor="F5F9F5") if alt else None
        data_align = Alignment(wrap_text=True, vertical="top")
        data_font = Font(name="Calibri", size=10)

        cell = ws.cell(row=row_i, column=1, value=fname)
        cell.alignment = data_align
        cell.font = data_font
        if fill:
            cell.fill = fill

        for col_i, f in enumerate(fields, start=2):
            cell = ws.cell(row=row_i, column=col_i, value=fields_map.get(f, ""))
            cell.alignment = data_align
            cell.font = data_font
            if fill:
                cell.fill = fill

    ws.freeze_panes = "B2"


def _build_metadata_sheet(ws, prompts: list[dict]) -> None:
    ws.title = "Metadata"
    rows = [
        ("Analysis Date", NOW.strftime("%Y-%m-%d %H:%M:%S")),
        ("Prompts Count", len(prompts)),
        ("Generator", "AIDOC — Python App"),
    ]
    bold = Font(bold=True, name="Calibri")
    normal = Font(name="Calibri")
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(i, 1, k).font = bold
        ws.cell(i, 2, str(v)).font = normal
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40


def _build_prompts_sheet(ws, prompts: list[dict]) -> None:
    ws.title = "Prompts Used"
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill("solid", fgColor=BRAND_GREEN)
    for col, h in enumerate(["field_name", "instruction"], 1):
        c = ws.cell(1, col, h)
        c.font = header_font
        c.fill = header_fill
    for i, p in enumerate(prompts, 2):
        ws.cell(i, 1, p.get("field_name", ""))
        ws.cell(i, 2, p.get("instruction", ""))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def write_pdf_report(results: list[tuple], prompts: list[dict]) -> bytes:
    fields, matrix = _pivot(results, prompts)
    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
        title="AIDOC — Análisis de Documentos",
    )

    styles = getSampleStyleSheet()
    story = []

    story += _cover_page(styles, len(matrix), len(fields))
    story.append(PageBreak())

    for doc_name, fields_map in matrix.items():
        story += _document_section(styles, doc_name, fields, fields_map)

    story += _summary_table(styles, fields, matrix)

    doc.build(story)
    return buf.getvalue()


def _cover_page(styles, doc_count: int, field_count: int) -> list:
    title_style = ParagraphStyle(
        "CoverTitle",
        fontSize=32,
        fontName="Helvetica-Bold",
        textColor=BRAND_GREEN_RGB,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    sub_style = ParagraphStyle(
        "CoverSub",
        fontSize=14,
        fontName="Helvetica",
        textColor=colors.HexColor("#4a4843"),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "CoverMeta",
        fontSize=11,
        fontName="Helvetica",
        textColor=colors.HexColor("#8a8780"),
        alignment=TA_CENTER,
    )

    return [
        Spacer(1, 4 * cm),
        Paragraph("AIDOC", title_style),
        Paragraph("Análisis Inteligente de Documentos", sub_style),
        HRFlowable(width="60%", color=BRAND_GREEN_RGB, thickness=2, spaceAfter=20),
        Paragraph(f"Documentos analizados: <b>{doc_count}</b>", meta_style),
        Paragraph(f"Campos evaluados: <b>{field_count}</b>", meta_style),
        Spacer(1, 0.4 * cm),
        Paragraph(f"Generado: {NOW.strftime('%d/%m/%Y %H:%M')}", meta_style),
    ]


def _document_section(styles, doc_name: str, fields: list[str], fields_map: dict) -> list:
    section_style = ParagraphStyle(
        "SectionTitle",
        fontSize=13,
        fontName="Helvetica-Bold",
        textColor=BRAND_GREEN_RGB,
        spaceBefore=16,
        spaceAfter=8,
    )
    elems = [
        Paragraph(doc_name, section_style),
        HRFlowable(width="100%", color=colors.HexColor("#e0ddd5"), thickness=1, spaceAfter=8),
    ]

    table_data = [["Campo", "Respuesta"]]
    for f in fields:
        table_data.append([f, fields_map.get(f, "—")])

    t = Table(table_data, colWidths=[4.5 * cm, 12 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN_RGB),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#eceae3")),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f9f5")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#c5d8c5")),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("WORDWRAP",   (0, 0), (-1, -1), True),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.5 * cm))
    return elems


def _summary_table(styles, fields: list[str], matrix: dict) -> list:
    title_style = ParagraphStyle(
        "SummaryTitle",
        fontSize=14,
        fontName="Helvetica-Bold",
        textColor=BRAND_GREEN_RGB,
        spaceBefore=20,
        spaceAfter=10,
    )
    elems = [
        PageBreak(),
        Paragraph("Resumen comparativo", title_style),
        HRFlowable(width="100%", color=BRAND_GREEN_RGB, thickness=1.5, spaceAfter=12),
    ]

    header = ["Documento"] + fields
    rows = [header]
    for fname, fm in matrix.items():
        short_name = fname[:30] + "…" if len(fname) > 30 else fname
        rows.append([short_name] + [fm.get(f, "—") for f in fields])

    col_w = [4 * cm] + [((A4[0] - 5 * cm - 4 * cm) / max(len(fields), 1))] * len(fields)
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN_RGB),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f9f5")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#c5d8c5")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(t)
    return elems


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — STREAMLIT APP (app.py)
# ══════════════════════════════════════════════════════════════════════════════

import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="AIDOC",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

if "messages" not in st.session_state:
    st.session_state.messages = []
if "results" not in st.session_state:
    st.session_state.results = []
if "prompts" not in st.session_state:
    st.session_state.prompts = []

# ─── Sidebar ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 📄 AIDOC")
    st.caption("Análisis Inteligente de Documentos")
    st.divider()

    st.markdown("**Configuración activa**")
    st.caption(f"LLM: `{settings.LLM_MODEL}`")
    st.caption(f"Embeddings: `{settings.EMBED_MODEL}`")
    st.caption(f"Qdrant: `{settings.QDRANT_URL}`")
    st.caption(f"Colección: `{settings.QDRANT_COLLECTION}`")

    st.divider()

    if st.button("🔧 Inicializar colección", use_container_width=True):
        with st.spinner("Conectando a Qdrant..."):
            try:
                ensure_collection()
                st.success("Colección lista")
            except Exception as e:
                st.error(f"Error: {e}")

    st.divider()

    try:
        docs = list_documents()
        st.metric("Documentos en Qdrant", len(docs))
        if docs:
            with st.expander(f"Ver {len(docs)} documento(s)"):
                for d in docs:
                    st.caption(f"📄 {d}")
    except Exception:
        st.warning("Qdrant no disponible")

    if st.session_state.results:
        st.divider()
        st.metric("Análisis completado", f"{len(st.session_state.results)} resp.")

# ─── Main ─────────────────────────────────────────────────────────────────────

st.title("📄 AIDOC")
st.markdown("Análisis inteligente de documentos — ingesta, análisis batch, exportación y chatbot.")

tab1, tab2, tab3, tab4 = st.tabs([
    "📂 1 · Ingestar",
    "⚡ 2 · Analizar",
    "📊 3 · Exportar",
    "💬 4 · Chatbot",
])

# ── Tab 1: Ingest ─────────────────────────────────────────────────────────────

with tab1:
    st.subheader("Cargar documentos")
    st.markdown("Sube PDFs, TXT o ZIPs. Los archivos se fragmentan, embedean y almacenan en Qdrant.")

    uploaded_files = st.file_uploader(
        "Archivos",
        type=["pdf", "txt", "zip"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded_files:
        st.caption(f"{len(uploaded_files)} archivo(s) seleccionado(s)")

    col_btn, col_note = st.columns([1, 3])
    with col_btn:
        ingest_btn = st.button("🚀 Ingestar", type="primary", disabled=not uploaded_files, use_container_width=True)

    if ingest_btn and uploaded_files:
        progress_bar = st.progress(0.0, text="Iniciando...")
        status = st.empty()
        errors: list[str] = []

        with tempfile.TemporaryDirectory() as tmpdir:
            for i, f in enumerate(uploaded_files):
                status.info(f"⏳ Procesando **{f.name}**...")
                tmp_path = Path(tmpdir) / f.name
                tmp_path.write_bytes(f.read())

                try:
                    if f.name.lower().endswith(".zip"):
                        ingest_zip(str(tmp_path))
                    else:
                        ingest_file(str(tmp_path), f.name)
                except Exception as e:
                    errors.append(f"{f.name}: {e}")

                progress_bar.progress((i + 1) / len(uploaded_files), text=f.name)

        status.empty()
        progress_bar.empty()

        if errors:
            for err in errors:
                st.error(f"✗ {err}")
        st.success(f"✅ {len(uploaded_files) - len(errors)} archivo(s) ingestados correctamente")
        st.rerun()

# ── Tab 2: Analyze ────────────────────────────────────────────────────────────

with tab2:
    st.subheader("Análisis batch")
    st.markdown("Selecciona un preset de prompts o sube tu propio CSV (`field_name,instruction`).")

    col_left, col_right = st.columns([1, 1])

    with col_left:
        preset = st.selectbox(
            "Preset",
            ["sample", "legal", "financial", "hr"],
            format_func=lambda x: {
                "sample": "Muestra general",
                "legal": "Contratos legales",
                "financial": "Due diligence financiero",
                "hr": "Contratos laborales",
            }[x],
        )
        preset_path = Path("prompts") / f"{preset}_prompts.csv"

    with col_right:
        custom_csv = st.file_uploader("O sube tu propio CSV", type=["csv"], key="custom_prompts")

    try:
        if custom_csv:
            prompts_df = pd.read_csv(custom_csv)
        else:
            prompts_df = pd.read_csv(preset_path)

        st.dataframe(prompts_df, use_container_width=True, hide_index=True)
    except FileNotFoundError:
        st.warning(f"No se encontró `{preset_path}`. Sube un CSV personalizado.")
        prompts_df = pd.DataFrame(columns=["field_name", "instruction"])

    st.divider()

    try:
        current_docs = list_documents()
    except Exception:
        current_docs = []

    if not current_docs:
        st.info("No hay documentos en Qdrant. Ve al Tab 1 para ingestar primero.")
    elif prompts_df.empty:
        st.warning("Define al menos un prompt.")
    else:
        n_docs = len(current_docs)
        n_prompts = len(prompts_df)
        st.info(f"Se ejecutarán **{n_docs} × {n_prompts} = {n_docs * n_prompts}** llamadas al LLM.")

        if st.button("⚡ Ejecutar análisis", type="primary"):
            prompts_list = prompts_df.to_dict("records")
            total = n_docs * n_prompts
            progress = st.progress(0.0, text="Iniciando análisis...")
            completed_tracker = {"n": 0}

            def on_progress(completed: int, total_: int, result: tuple):
                completed_tracker["n"] = completed
                pct = completed / total_
                fname, field, _ = result
                progress.progress(pct, text=f"[{completed}/{total_}] {fname} · {field}")

            try:
                results = run_batch_analysis(current_docs, prompts_list, on_progress)
                st.session_state.results = results
                st.session_state.prompts = prompts_list
                progress.empty()
                st.success(f"✅ Análisis completo — {len(results)} respuestas generadas")
            except Exception as e:
                progress.empty()
                st.error(f"Error durante el análisis: {e}")
                results = []

            if results:
                matrix: dict = {}
                for fname, field, answer in results:
                    matrix.setdefault(fname, {"Filename": fname})[field] = answer
                df_preview = pd.DataFrame(matrix.values())
                st.dataframe(df_preview, use_container_width=True)

# ── Tab 3: Export ─────────────────────────────────────────────────────────────

with tab3:
    st.subheader("Exportar resultados")

    if not st.session_state.results:
        st.info("Ejecuta el análisis (Tab 2) para habilitar la exportación.")
    else:
        n_res = len(st.session_state.results)
        n_docs_res = len({r[0] for r in st.session_state.results})
        st.success(f"Resultados disponibles: **{n_docs_res}** documentos · **{n_res}** respuestas")

        col_xl, col_pdf = st.columns(2)

        with col_xl:
            st.markdown("### 📊 Excel")
            st.caption("Matriz completa — una fila por documento, una columna por campo.")
            with st.spinner("Generando Excel..."):
                excel_bytes = write_excel(st.session_state.results, st.session_state.prompts)
            st.download_button(
                "⬇️ Descargar Excel",
                data=excel_bytes,
                file_name="aidoc_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        with col_pdf:
            st.markdown("### 📋 PDF Report")
            st.caption("Reporte estructurado — portada, sección por documento, tabla resumen.")
            with st.spinner("Generando PDF..."):
                pdf_bytes = write_pdf_report(st.session_state.results, st.session_state.prompts)
            st.download_button(
                "⬇️ Descargar PDF",
                data=pdf_bytes,
                file_name="aidoc_report.pdf",
                mime="application/pdf",
                use_container_width=True,
                type="primary",
            )

# ── Tab 4: Chatbot ────────────────────────────────────────────────────────────

with tab4:
    st.subheader("Chatbot conversacional")
    st.caption("Pregunta sobre los documentos analizados. Usa RAG sobre Qdrant + la matriz de resultados.")

    if not st.session_state.messages:
        with st.chat_message("assistant"):
            st.markdown(
                "Hola, soy el asistente AIDOC. Puedo ayudarte a analizar los resultados de los documentos procesados. "
                "Puedo comparar documentos, explicar campos específicos o profundizar en el contenido original. "
                "¿Qué quieres explorar?"
            )

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    if user_input := st.chat_input("Pregunta sobre los documentos..."):
        st.session_state.messages.append({"role": "user", "content": user_input})
        with st.chat_message("user"):
            st.markdown(user_input)

        with st.chat_message("assistant"):
            with st.spinner("Consultando documentos..."):
                try:
                    response = chat(
                        st.session_state.messages,
                        st.session_state.results or None,
                    )
                except Exception as e:
                    response = f"Error al consultar el LLM: {e}"
            st.markdown(response)

        st.session_state.messages.append({"role": "assistant", "content": response})

    if st.session_state.messages:
        if st.button("🗑️ Limpiar conversación", use_container_width=False):
            st.session_state.messages = []
            st.rerun()
