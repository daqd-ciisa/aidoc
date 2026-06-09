"""Parser de hojas de cálculo: ``.xlsx`` (openpyxl) y ``.csv`` (stdlib).

Renderiza cada hoja como filas ``celda | celda | …`` (igual que las tablas de docx),
para que un catálogo de servicios/productos en Excel quede indexable y utilizable en el
chat y las cotizaciones. Sin paginación (``page=None``)."""
from __future__ import annotations

import csv
import logging

from openpyxl import load_workbook

from app.ingestion.parsers.base import ParsedPage

logger = logging.getLogger("aidoc.parsers.excel")


def _cell(value: object) -> str:
    """Texto de una celda (vacío si None), en una sola línea."""
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace("\r", " ")


def _row_text(values) -> str | None:
    """Renderiza una fila como ``celda | celda | …`` (recortando celdas vacías al final).

    Devuelve ``None`` si la fila está totalmente vacía."""
    cells = [_cell(v) for v in values]
    while cells and not cells[-1]:
        cells.pop()
    if not cells:
        return None
    return " | ".join(cells)


def parse_xlsx(path: str) -> list[ParsedPage]:
    """Lee un .xlsx: cada hoja se vuelca como filas, encabezada por su nombre.

    ``data_only=True`` toma el valor calculado de las fórmulas (no la fórmula)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for ws in wb.worksheets:
            rows = [t for t in (_row_text(r) for r in ws.iter_rows(values_only=True)) if t]
            if not rows:
                continue
            title = (ws.title or "").strip()
            block = "\n".join(rows)
            parts.append(f"Hoja: {title}\n{block}" if title else block)
    finally:
        wb.close()
    full = "\n\n".join(parts).strip()
    return [ParsedPage(text=full, page=None)] if full else []


def parse_csv(path: str) -> list[ParsedPage]:
    """Lee un .csv tolerando encoding (BOM) y dialecto (coma/punto y coma)."""
    rows: list[str] = []
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect: type[csv.Dialect] | csv.Dialect = (
                csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else csv.excel
            )
        except csv.Error:
            dialect = csv.excel
        for row in csv.reader(fh, dialect):
            text = _row_text(row)
            if text:
                rows.append(text)
    full = "\n".join(rows).strip()
    return [ParsedPage(text=full, page=None)] if full else []
